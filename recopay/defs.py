import time

from openpyxl import load_workbook, Workbook
import pandas as pd
import traceback
from datetime import datetime as date
from keyboards import *
from config import *
from const import *
from pymysql import connect, OperationalError
import phonenumbers


########################################################################################################################
def error(err):
    try:
        import os
        file=open("log_error.txt", "a", encoding='utf-8')
        file.write(f"\n\n###########{err}\n"
                   f"\n{date.now()}\n###########\n\n")
        file.close()
    except Exception:
        print(traceback.format_exc())
########################################################################################################################
#########################################_ФУНКЦИИ БАЗЫ ДАННЫХ_#######################################################
def close_connection(db, sql):
    try:
        sql.close()
        db.close()
    except OperationalError as e:
        print()
        print('####################################')
        print(f'Ошибка создания соединения\n\n{e}')
        print(f'\n{date.now()}')
        print('####################################')


def start_connection():
    try:
        db=connect(
            host=ip_address_db,
            port=port_db,
            user=login_db,
            password=password_db,
            database=name_db,
            autocommit=True
        )
        sql=db.cursor()
        return db, sql
    except Exception:
        error(traceback.format_exc())
########################################################################################################################
########################################################################################################################
def create_WaitPartner():
    try:
        db, sql = start_connection()
        sql.execute("""create table if not exists `WaitPartner`(
                            `ID` int(11) NOT NULL AUTO_INCREMENT Primary Key,
                            `AdminID` bigint,
                            `LoginTG` varchar(50),
                            `FullName` varchar(155),
                            `Sum` int(11)
                    )""")
        close_connection(db, sql)
    except Exception:
        error(traceback.format_exc())
########################################################################################################################
########################################################################################################################
def create_WaitForSend():
    try:
        db, sql = start_connection()
        sql.execute(f"""create TABLE if not EXISTS `WaitForSend`(
            `ID` int(11) NOT NULL AUTO_INCREMENT Primary Key,
            `PartnerID` int(11) not null,
            `ClientID` int(11) not null
            )""")
        close_connection(db, sql)
    except Exception:
        error(traceback.format_exc())
########################################################################################################################
########################################################################################################################
def create_WaitClient():
    try:
        db, sql = start_connection()
        sql.execute("""create table if not exists `WaitClient`(
                            `ID` int(11) NOT NULL AUTO_INCREMENT Primary Key,
                            `TelegramID` bigint,
                            `LastName` varchar(50),
                            `FirstName` varchar(50),
                            `FatherName` varchar(50),
                            `Phone` varchar(12)
                    )""")
        close_connection(db, sql)
    except Exception:
        error(traceback.format_exc())
########################################################################################################################
########################################################################################################################
def create_PartnerTelegram():
    try:
        db, sql = start_connection()
        sql.execute("""create table if not exists `PartnerTelegram`(
                            `ID` int(11) NOT NULL AUTO_INCREMENT Primary Key,
                            `PartnerID` int(11) not null,
                            `TelegramID` bigint
                    )""")
        #sql.execute("ALTER TABLE `PartnerTelegram` ADD CONSTRAINT fk_partnerID_id FOREIGN KEY (`PartnerID`) REFERENCES `Partner`(`ID`);")
        close_connection(db, sql)
    except Exception:
        error(traceback.format_exc())
########################################################################################################################
########################################################################################################################
def create_WaitData():
    try:
        db, sql = start_connection()
        sql.execute("""create table if not exists `WaitData`(
                            `ID` int(11) NOT NULL AUTO_INCREMENT Primary Key,
                            `AdminID` bigint,
                            `ValueData` varchar(155)
                    )""")
        close_connection(db, sql)
    except Exception:
        error(traceback.format_exc())
########################################################################################################################
########################################################################################################################
def get_data_from_xlsx(file_name, bot):
    try:

        client_data=[]
        name_lists=pd.ExcelFile(file_name).sheet_names
        for i in range(len(name_lists)):
            data=pd.DataFrame(pd.read_excel(file_name, sheet_name=name_lists[i]))
            target=data.columns.values
            result=""
            print(data)
            for z in range(len(target)):
                if z == len(target) - 1:
                    result+=f"{target[z]}"
                    break
                result+=f"{target[z]}?"
            if result != "":
                client_data.append(result)
            for z in range(len(data)):
                result=""
                values=data[target].loc[data.index[z]]
                x=str(values).split(' ')
                while "" in x: x.remove("")
                x=" ".join(x).split("\n")
                del x[-1]
                for k in range(len(x)):
                    wait=x[k].split(' ')
                    for e in range(len(target)):
                        if str(target[e]) in wait:
                            wait.remove(str(target[e]))
                            if e == len(target) - 1:
                                result+=f"{wait[0]}"
                                break
                            result+=f"{wait[0]}?"
                            break
                client_data.append(result)
        print(client_data)
        db, sql = start_connection()
        sql.execute(f"""truncate `PotentialClient`;""")
        sql.execute(f"""truncate `WaitForSend`;""")
        if len(client_data) > 0:
            for i in range(len(client_data)):
                x=client_data[i].split('?')
                try:
                    phone=phonenumbers.format_number(phonenumbers.parse(x[-1],'RU'),phonenumbers.PhoneNumberFormat.E164)
                except Exception:
                    phone = "-"
                del x[-1]
                sql.execute("select count(*) from `Client` where (`Surname` = %s or `Name` = %s or `LastName` = %s) "
                            "and `IsPayment` = 0",
                            (x[0], x[1], x[2],))
                if sql.fetchone()[0] == 0:
                    pass
                else:
                    sql.execute(
                    "insert into `PotentialClient`(`Surname`, `Name`, `LastName`, `Telephone`) values(%s, %s, %s, %s)",
                        (x[0], x[1], x[2], phone))
                    sql.execute("select count(*) from `Client` where (`Telephone` = %s and `Surname` = %s and "
                                        "`LastName`= %s) or (`Telephone` = %s and `Name` = %s and `LastName`= %s)"
                                " and `IsPayment` = 0",
                                (phone, x[0], x[2], phone, x[1], x[2],))
                    if sql.fetchone()[0] == 0:
                        pass
                    else:
                        sql.execute("insert into `PotentialClient`(`Surname`, `Name`, `LastName`, `Telephone`) "
                                    "values(%s, %s, %s, %s)",
                            (x[0], x[1], x[2], phone))
        dd, ss = start_connection()
        sql.execute("select * from `PotentialClient` order by `ID` desc")
        while True:
            row = sql.fetchone()
            if row is None:
                break
            ss.execute(
                        "select count(*) from `Client` where `IsPayment` = 1 and `Surname` ="
                        " %s and `Name` = %s and `LastName` = %s and `Telephone` = %s",
                       (row[1], row[2], row[3], row[4],))
            if ss.fetchone()[0] > 0:
                print("delete")
                ss.execute("delete from `PotentialClient` where `ID` = %s", row[0])
        sql.execute("select * from `PotentialClient` order by `ID`")
        while True:
            row = sql.fetchone()
            if row is None:
                break
            ss.execute(
"select count(*) from `PotentialClient` where `Surname` = %s and `Name` = %s and `LastName` = %s and `Telephone` = %s",
                       (row[1], row[2], row[3], row[4],))
            if ss.fetchone()[0] > 1:
                print("delete")
                ss.execute(
"delete from `PotentialClient` where `Surname` = %s and `Name` = %s and `LastName` = %s and `Telephone` = %s and `ID` != %s",
                    (row[1], row[2], row[3], row[4], row[0], ))
        id_data=[]
        db, sql=start_connection()
        dd, ss=start_connection()
        sql.execute("select * from `Client` order by id asc;")
        while True:
            row=sql.fetchone()
            if row is None:
                break
            if row[0] not in id_data:
                # ss.execute(f"""select count(*) from `Client` where `IsPayment` = 0 and `ID` != %(id)s and
                #                 ((`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s)
                #                 or
                #                 (`Telephone` = %(phone)s and `Name` = %(name)s and `LastName` = %(lastname)s)
                #                 or                                                                                                 ВЕРСИЯ С НОМЕРОМ
                #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s and
                #                 `Name` = %(name)s)
                #                 or
                #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `Name` = %(name)s));""",
                #            {"id": int(row[0]), "name": row[2], "surname": row[1], "lastname": row[3], "phone": row[4]})
                ss.execute(f"""select count(*) from `Client` where `IsPayment` = 0 and `ID` != %(id)s and 
                                ((`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s)
                                or
                                (`Telephone` = %(phone)s and `Name` = %(name)s and `LastName` = %(lastname)s) 
                                or 
                                (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s and 
                                `Name` = %(name)s)
                                or
                                (`Telephone` = %(phone)s and `Surname` = %(surname)s and `Name` = %(name)s)      
                                or
                                (`Surname` = %(surname)s and `LastName` = %(lastname)s and `Name` = %(name)s));""",
                           {"id": int(row[0]), "name": row[2], "surname": row[1], "lastname": row[3], "phone": row[4]})
                if ss.fetchone()[0] > 0:
                    # ss.execute(f"""select * from `Client` where `IsPayment` = 0 and `ID` != %(id)s and
                    #                 ((`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s)
                    #                 or
                    #                 (`Telephone` = %(phone)s and `Name` = %(name)s and `LastName` = %(lastname)s)
                    #                 or                                                                                            ВЕРСИЯ С НОМЕРОМ
                    #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s and
                    #                 `Name` = %(name)s)
                    #                 or
                    #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `Name` = %(name)s));""",
                    #            {"id": int(row[0]), "name": row[2], "surname": row[1], "lastname": row[3],
                    #             "phone": row[4]})
                    ss.execute(f"""select * from `Client` where `IsPayment` = 0 and `ID` != %(id)s and
                                    ((`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s)
                                    or
                                    (`Telephone` = %(phone)s and `Name` = %(name)s and `LastName` = %(lastname)s)
                                    or
                                    (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s and
                                    `Name` = %(name)s)
                                    or
                                    (`Telephone` = %(phone)s and `Surname` = %(surname)s and `Name` = %(name)s)
                                                                or
                                    (`Surname` = %(surname)s and `LastName` = %(lastname)s and `Name` = %(name)s));""",
                               {"id": int(row[0]), "name": row[2], "surname": row[1], "lastname": row[3],
                                "phone": row[4]})
                    while True:
                        riw=ss.fetchone()
                        if riw is None:
                            break
                        id_data.append(riw[0])
        partner=[]
        sql.execute(f"""truncate `WaitForSend`;""")
        query = f"""select * from `Client` where `ID` not in %(data)s and `IsPayment` = 0 order by id;"""
        if len(id_data) == 0:
            query = f"""select * from `Client` where `IsPayment` = 0 order by id;"""
        sql.execute(query, {"data": tuple(id_data)})
        while True:
            row=sql.fetchone()
            if row is None:
                break

            # ss.execute(f"""select count(*) from `PotentialClient` where
            #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s)
            #                 or
            #                 (`Telephone` = %(phone)s and `Name` = %(name)s and `LastName` = %(lastname)s)
            #                 or                                                                                         С НОМЕРОМ
            #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s and
            #                 `Name` = %(name)s)
            #                 or
            #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `Name` = %(name)s);""",
            #            {"name": row[2], "surname": row[1], "lastname": row[3], "phone": row[4]})
            query_count = f"""select count(*) from `PotentialClient` where 
                            (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s)
                            or
                            (`Telephone` = %(phone)s and `Name` = %(name)s and `LastName` = %(lastname)s) 
                            or 
                            (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s and 
                            `Name` = %(name)s)
                            or
                            (`Telephone` = %(phone)s and `Surname` = %(surname)s and `Name` = %(name)s)
                                or
                            (`Surname` = %(surname)s and `LastName` = %(lastname)s and `Name` = %(name)s);"""
            ss.execute(query_count,
                       {"name": row[2], "surname": row[1], "lastname": row[3], "phone": row[4]})

            if ss.fetchone()[0] > 0:
                # ss.execute(f"""delete from `PotentialClient` where
                #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s)
                #                 or
                #                 (`Telephone` = %(phone)s and `Name` = %(name)s and `LastName` = %(lastname)s)
                #                 or                                                                                        С НОМЕРОМ
                #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s and
                #                 `Name` = %(name)s)
                #                 or
                #                 (`Telephone` = %(phone)s and `Surname` = %(surname)s and `Name` = %(name)s);""",
                #            {"name": row[2], "surname": row[1], "lastname": row[3], "phone": row[4]})
                querys_select = f"""delete from `PotentialClient` where 
                                (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s)
                                or
                                (`Telephone` = %(phone)s and `Name` = %(name)s and `LastName` = %(lastname)s) 
                                or 
                                (`Telephone` = %(phone)s and `Surname` = %(surname)s and `LastName` = %(lastname)s and 
                                `Name` = %(name)s)
                                or
                                (`Telephone` = %(phone)s and `Surname` = %(surname)s and `Name` = %(name)s)
                                    or
                                (`Surname` = %(surname)s and `LastName` = %(lastname)s and `Name` = %(name)s);"""
                ss.execute(querys_select,
                           {"name": row[2], "surname": row[1], "lastname": row[3], "phone": row[4]})

                ss.execute(f"""select * from `ClientPartner` where `ClientID` = %(id)s;""", {"id": row[0]})
                riw=ss.fetchone()
                ss.execute(f"""insert into `WaitForSend`(`PartnerID`, `ClientID`) values(%(id)s, %(ids)s);""",
                           {"id": riw[1], "ids": row[0]})
                if riw[1] not in partner:
                    partner.append(riw[1])
        if partner is None:
            close_connection(db, sql)
            close_connection(dd, ss)
            return None
        final_data=[]
        for i in range(len(partner)):
            sql.execute(f"""select * from `Partner` where `ID` = %(id)s;""", {"id": partner[i]})
            riw=sql.fetchone()
            sql.execute(f"""select * from `PartnerTelegram` where `PartnerID` = %(id)s;""", {"id": partner[i]})
            ids=sql.fetchone()
            coun=1
            data_msg=[]
            msg="Клиенты, которые пришли в клинику по вашей рекомендации:\n\n"
            sql.execute(f"""select * from `WaitForSend` where `PartnerID` = %(id)s;""", {"id": partner[i]})
            while True:
                row=sql.fetchone()
                if row is None:
                    xx = f"\nОбщая сумма на оплату: {(coun-1)*riw[-1]}"
                    if (len(xx) + len(msg)) < 1500:
                        msg+=xx
                        break
                    else:
                        data_msg.append(msg)
                        msg=xx
                        break
                ss.execute(f"""select * from `Client` where `ID` = %(id)s;""", {"id": row[-1]})
                datas=ss.fetchone()
                ss.execute(f"""update `Client` set `IsPayment` = 1 where `ID` = %(id)s;""", {"id": row[-1]})
                wait=f"{coun}. {datas[1]} {datas[2]} {datas[3]}\n"
                coun+=1
                if (len(wait) + len(msg)) < 1500:
                    msg+=wait
                else:
                    data_msg.append(msg)
                    msg=wait
            if msg != "":
                data_msg.append(msg)
            final_data.append(f"{riw[2]}?{riw[-1] * (coun - 1)}")
            print(final_data)
            for z in range(len(data_msg)):
                print(data_msg[z])
                bot.send_message(ids[-1], data_msg[z])
                time.sleep(0.4)
        close_connection(db, sql)
        close_connection(dd, ss)
        summary=0
        filename = ""
        print(final_data)
        fn=Workbook()
        fn.save("create_data.xlsx")
        fn.close()
        fn="create_data.xlsx"
        ws=load_workbook(fn)
        wb=ws["Sheet"]
        for i in range(len(final_data)):
            x=final_data[i].split('?')
            summary+=int(x[-1])
            wb.append([x[0], int(x[-1])])
            if i == len(final_data) - 1:
                wb.append(["ИТОГО:", summary])
                filename = f"Отчет для администратора {date.now().date()}.xlsx"
                ws.save(filename)
                ws.close()
        return filename
        # query = "select `ID`, `FullName`, (select count(*) from `ClientPartner` where `PartnerID` = `Partner`.`ID` and " \
        #         "`ClientID` in (select`Client`.`ID` from `Client` JOIN `PotentialClient` WHERE (`Client`.`Surname` = " \
        #         "`PotentialClient`.`Surname` and `Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = " \
        #         "`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #         "OR" \
        #         " (`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`LastName` = " \
        #         "`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`)" \
        #         " OR " \
        #         "(`Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = `PotentialClient`.`LastName` " \
        #         "AND `Client`.`Telephone` = `PotentialClient`.`Telephone`)))*`Sum` as `Total` from `Partner` where" \
        #         " (select count(*) from `ClientPartner` where `PartnerID` = `Partner`.`ID` and `ClientID` in " \
        #         "(select `Client`.`ID` from `Client` JOIN `PotentialClient` WHERE (`Client`.`Surname` = " \
        #         "`PotentialClient`.`Surname` and `Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = " \
        #         "`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #         "OR" \
        #         " (`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`LastName` = " \
        #         "`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #         "OR" \
        #         " (`Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = `PotentialClient`.`LastName` " \
        #         "AND `Client`.`Telephone` = `PotentialClient`.`Telephone`)) > 0)"
        # c = 1
        # summary = 0
        # fn = Workbook()
        # fn.save("create_data.xlsx")
        # fn.close()
        # fn = "create_data.xlsx"
        # ws = load_workbook(fn)
        # wb = ws["Sheet"]
        # data_id_partner = []
        # sql.execute(query)
        # while True:
        #     riw = sql.fetchone()
        #     print()
        #     print(riw)
        #     if riw is None:
        #         break
        #     if riw[0] not in data_id_partner:
        #         data_id_partner.append(riw[0])
        #     summary += int(riw[2])
        #     wb.append([riw[1], int(riw[2])])
        #     c +=1
        # if c == 1:
        #     ws.close()
        #     close_connection(db, sql)
        #     return None
        # for i in range(len(data_id_partner)):
        #     msg = "Клиенты, которые пришли в клинику по вашей рекомендации:\n\n"
        #     wait = ""
        #     data_msg = []
        #     c = 1
        #     quer="select count(*) from `Client` where `ID` in (select `ClientID` from `ClientPartner` where " \
        #          "`ClientPartner`.`PartnerID` = %s and `ClientID` in(select `Client`.`ID` from `Client` JOIN " \
        #          "`PotentialClient` WHERE (`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`Name` = " \
        #          "`PotentialClient`.`Name` and `Client`.`LastName` =  `PotentialClient`.`LastName` AND " \
        #          "`Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #          "OR" \
        #          " (`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`LastName` = " \
        #          "`PotentialClient`.`LastName` AND  `Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #          "OR" \
        #          " (`Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = `PotentialClient`.`LastName`" \
        #          " AND `Client`.`Telephone` = `PotentialClient`.`Telephone`))) group by `Telephone`"
        #     sql.execute(quer, data_id_partner[i])
        #     if sql.fetchone()[0] > 0:
        #         quer="select * from `Client` where `ID` in (select `ClientID` from `ClientPartner` where " \
        #              "`ClientPartner`.`PartnerID` = %s and `ClientID` in(select `Client`.`ID` from `Client` JOIN " \
        #              "`PotentialClient` WHERE (`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`Name` = " \
        #              "`PotentialClient`.`Name` and `Client`.`LastName` =  `PotentialClient`.`LastName` AND " \
        #              "`Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #              "OR" \
        #              " (`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`LastName` = " \
        #              "`PotentialClient`.`LastName` AND  `Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #              "OR" \
        #              " (`Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = `PotentialClient`.`LastName`" \
        #              " AND `Client`.`Telephone` = `PotentialClient`.`Telephone`))) group by `Telephone`"
        #         dd, ss = start_connection()
        #         sql.execute(quer, data_id_partner[i])
        #         while True:
        #             row = sql.fetchone()
        #             if row is None:
        #                 break
        #             ss.execute("update `Client` set `IsPayment` = 1 where `ID` = %s", row[0])
        #             wait = f"{c}. {row[1]} {row[2]} {row[3]}\n"
        #             if len(wait) + len(msg) <= 1024:
        #                 msg+=wait
        #             else:
        #                 data_msg.append(msg)
        #                 msg=wait
        #             c += 1
        #         query="select `ID`, `FullName`, (select count(*) from `ClientPartner` where `PartnerID` = `Partner`.`ID` and " \
        #               "`ClientID` in (select`Client`.`ID` from `Client` JOIN `PotentialClient` WHERE (`Client`.`Surname` = " \
        #               "`PotentialClient`.`Surname` and `Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = " \
        #               "`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #               "OR" \
        #               " (`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`LastName` = " \
        #               "`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`)" \
        #               " OR " \
        #               "(`Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = `PotentialClient`.`LastName` " \
        #               "AND `Client`.`Telephone` = `PotentialClient`.`Telephone`)))*`Sum` as `Total` from `Partner` where" \
        #               " (select count(*) from `ClientPartner` where `PartnerID` = `Partner`.`ID` and `ClientID` in " \
        #               "(select `Client`.`ID` from `Client` JOIN `PotentialClient` WHERE (`Client`.`Surname` = " \
        #               "`PotentialClient`.`Surname` and `Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = " \
        #               "`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #               "OR" \
        #               " (`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`LastName` = " \
        #               "`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`) " \
        #               "OR" \
        #               " (`Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = `PotentialClient`.`LastName` " \
        #               "AND `Client`.`Telephone` = `PotentialClient`.`Telephone`)) > 0 and `ID` = %s)"
        #         sql.execute(query, data_id_partner[i])
        #         while True:
        #             row = sql.fetchone()
        #             if row is None:
        #                 break
        #             if row[0] == data_id_partner[i]:
        #                 wait = f"\nОбщая сумма на оплату: {row[-1]}"
        #                 if len(wait) + len(msg) <= 1024:
        #                     msg+=wait
        #                     data_msg.append(msg)
        #                     break
        #                 else:
        #                     data_msg.append(msg)
        #                     msg=wait
        #                     data_msg.append(msg)
        #                     break
        #         sql.execute("select * from `PartnerTelegram` where `PartnerID` = %s", data_id_partner[i])
        #         row = sql.fetchone()
        #         for z in range(len(data_msg)):
        #             bot.send_message(row[-1], data_msg[z])
        #             time.sleep(0.4)
        # sql.execute("truncate `PotentialClient`")
        # close_connection(db, sql)
        # close_connection(dd, ss)
        # wb.append(['Итого', summary])
        # ws.save(f"Отчет для администратора {date.now().date()}.xlsx")
        # ws.close()
        # return f"Отчет для администратора {date.now().date()}.xlsx"
    except Exception:
        error(traceback.format_exc())
        return None
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
"""
ИТОГОВЫЙ ЗАПРОС ДЛЯ ПОЛУЧЕНИЯ КЛИЕНТОВ ПО ПАРТНЕРУ

select * from `Client` where `ID` in (select `ClientID` from `ClientPartner` where
  `ClientPartner`.`PartnerID` = 2 and `ClientID` in(select `Client`.`ID` from `Client` JOIN `PotentialClient` 
 WHERE `Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`Name` = `PotentialClient`.`Name` and 
 `Client`.`LastName` = `PotentialClient`.`LastName` 
 AND `Client`.`Telephone` = `PotentialClient`.`Telephone`)) group by `Telephone`
"""
########################################################################################################################
########################################################################################################################
########################################################################################################################
"""
ИТОГОВЫЙ КОД ДЛЯ ЦЕНЫ
select count(*) from `Client` where `ID` in 
(select `ClientID` from `ClientPartner` where `ClientPartner`.`PartnerID` = 9 and `ClientID` in(select 
`Client`.`ID` from `Client` JOIN `PotentialClient` WHERE (`Client`.`Surname` = 
`PotentialClient`.`Surname` and `Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = 
`PotentialClient`.`LastName` AND `Client`.`Telephone` = `PotentialClient`.`Telephone`) OR 
(`Client`.`Surname` = `PotentialClient`.`Surname` and `Client`.`LastName` = `PotentialClient`.`LastName` AND 
`Client`.`Telephone` = `PotentialClient`.`Telephone`)
OR                                                                                                                                                                        
(`Client`.`Name` = `PotentialClient`.`Name` and `Client`.`LastName` = `PotentialClient`.`LastName` AND 
`Client`.`Telephone` = `PotentialClient`.`Telephone`))) group by `Telephone`
"""
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
